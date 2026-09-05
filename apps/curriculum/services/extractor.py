"""Extraction service for Cambridge curriculum frameworks from source syllabus workbooks and PDFs."""

from __future__ import annotations

import glob
import os
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
import pypdf


LO_PATTERN = re.compile(r"^(\*?[A-Za-z0-9\.\-]+[0-9]+[A-Za-z0-9\.\-]*)\s+(.*)", re.DOTALL)
STAGE_PATTERN = re.compile(r"Stage\s*([0-9]+(?:\s*(?:to|and|\-|&)\s*[0-9]+)?)", re.IGNORECASE)


SUBJECT_METADATA: Dict[str, Dict[str, Any]] = {
    # Cambridge Primary
    "0058": {"name": "English", "framework": "Cambridge Primary", "stages": ["Stage 1", "Stage 2", "Stage 3", "Stage 4", "Stage 5", "Stage 6"]},
    "0059": {"name": "Computing", "framework": "Cambridge Primary", "stages": ["Stage 1", "Stage 2", "Stage 3", "Stage 4", "Stage 5", "Stage 6"]},
    "0065": {"name": "Humanities", "framework": "Cambridge Primary", "stages": ["Stages 1 to 3", "Stages 4 to 6"]},
    "0067": {"name": "Art & Design", "framework": "Cambridge Primary", "stages": ["Stages 1 to 6"]},
    "0072": {"name": "Digital Literacy", "framework": "Cambridge Primary", "stages": ["Stage 1", "Stage 2", "Stage 3", "Stage 4", "Stage 5", "Stage 6"]},
    "0838": {"name": "Global Perspectives", "framework": "Cambridge Primary", "stages": ["Stage 1", "Stage 2", "Stages 3 and 4", "Stages 5 and 6"]},
    # Cambridge Lower Secondary
    "0073": {"name": "Art & Design", "framework": "Cambridge Lower Secondary", "stages": ["Stages 7 to 9"]},
    "0860": {"name": "Computing", "framework": "Cambridge Lower Secondary", "stages": ["Stage 7", "Stage 8", "Stage 9"]},
    "0082": {"name": "Digital Literacy", "framework": "Cambridge Lower Secondary", "stages": ["Stage 7", "Stage 8", "Stage 9"]},
    "0861": {"name": "English", "framework": "Cambridge Lower Secondary", "stages": ["Stage 7", "Stage 8", "Stage 9"]},
    "0876": {"name": "English as a Second Language", "framework": "Cambridge Lower Secondary", "stages": ["Stage 7", "Stage 8", "Stage 9"]},
    "1129": {"name": "Global Perspectives", "framework": "Cambridge Lower Secondary", "stages": ["Stages 7 and 8", "Stage 9"]},
    "0839": {"name": "Humanities", "framework": "Cambridge Lower Secondary", "stages": ["Stages 7 to 9"]},
    "0862": {"name": "Mathematics", "framework": "Cambridge Lower Secondary", "stages": ["Stage 7", "Stage 8", "Stage 9"]},
    "0078": {"name": "Music", "framework": "Cambridge Lower Secondary", "stages": ["Stages 7&8", "Stage 9"]},
    "0081": {"name": "Physical Education", "framework": "Cambridge Lower Secondary", "stages": ["Stages 7 to 9"]},
    "0893": {"name": "Science", "framework": "Cambridge Lower Secondary", "stages": ["Stage 7", "Stage 8", "Stage 9"]},
    "0859": {"name": "Wellbeing", "framework": "Cambridge Lower Secondary", "stages": ["Stages 7 to 9"]},
    # Cambridge IGCSE
    "0417": {"name": "Information and Communication Technology", "framework": "Cambridge IGCSE", "stages": ["Years 10-11 (IGCSE)"]},
    "0450": {"name": "Business Studies", "framework": "Cambridge IGCSE", "stages": ["Years 10-11 (IGCSE)"]},
    "0452": {"name": "Accounting", "framework": "Cambridge IGCSE", "stages": ["Years 10-11 (IGCSE)"]},
    "0455": {"name": "Economics", "framework": "Cambridge IGCSE", "stages": ["Years 10-11 (IGCSE)"]},
    "0457": {"name": "Global Perspectives", "framework": "Cambridge IGCSE", "stages": ["Years 10-11 (IGCSE)"]},
    "0460": {"name": "Geography", "framework": "Cambridge IGCSE", "stages": ["Years 10-11 (IGCSE)"]},
    "0470": {"name": "History", "framework": "Cambridge IGCSE", "stages": ["Years 10-11 (IGCSE)"]},
    "0478": {"name": "Computer Science", "framework": "Cambridge IGCSE", "stages": ["Years 10-11 (IGCSE)"]},
    "0500": {"name": "English - First Language", "framework": "Cambridge IGCSE", "stages": ["Years 10-11 (IGCSE)"]},
    "0511": {"name": "English as a Second Language", "framework": "Cambridge IGCSE", "stages": ["Years 10-11 (IGCSE)"]},
    "0580": {"name": "Mathematics", "framework": "Cambridge IGCSE", "stages": ["Years 10-11 (IGCSE)"]},
    "0610": {"name": "Biology", "framework": "Cambridge IGCSE", "stages": ["Years 10-11 (IGCSE)"]},
    "0620": {"name": "Chemistry", "framework": "Cambridge IGCSE", "stages": ["Years 10-11 (IGCSE)"]},
    "0625": {"name": "Physics", "framework": "Cambridge IGCSE", "stages": ["Years 10-11 (IGCSE)"]},
}


def clean_text(text: Any) -> str:
    if text is None:
        return ""
    val = str(text).strip()
    val = re.sub(r"\s+", " ", val)
    return val


def parse_lo_line(text: str) -> Optional[Tuple[str, str]]:
    text = clean_text(text)
    if not text:
        return None
    # Skip non-LO titles or table headers
    if text.lower().startswith("learning objective") or text.lower().startswith("sub-strand") or text.lower().startswith("strand"):
        return None
    match = LO_PATTERN.match(text)
    if match:
        code = match.group(1).strip()
        lo_text = match.group(2).strip()
        return code, lo_text
    return None


class CurriculumExtractor:
    """Extracts curriculum frameworks, subjects, schemes, topics, subtopics, and LOs."""

    def __init__(self, source_dir: str):
        self.source_dir = Path(source_dir)

    def discover_workbooks(self) -> List[Path]:
        files = glob.glob(str(self.source_dir / "**" / "*.xlsx"), recursive=True)
        return [Path(f) for f in sorted(files)]

    def discover_pdfs(self) -> List[Path]:
        files = glob.glob(str(self.source_dir / "**" / "*.pdf"), recursive=True)
        igcse_files = [Path(f) for f in sorted(files) if "igcse" in str(f).lower() or re.search(r"\b0[4-6]\d{2}\b", Path(f).name)]
        return igcse_files

    def extract_subject_info(self, file_path: Path) -> Tuple[str, str, str]:
        filename = file_path.stem
        code_match = re.search(r"\b(\d{4})\b", filename)
        code = code_match.group(1) if code_match else "0000"

        if code in SUBJECT_METADATA:
            meta = SUBJECT_METADATA[code]
            return code, meta["name"], meta["framework"]

        # Fallback detection
        name = filename.replace("Learning Objectives Only", "").replace("Learning Objectives", "").strip()
        if int(code) < 700:
            framework = "Cambridge Primary"
        elif int(code) < 1000:
            framework = "Cambridge Lower Secondary"
        else:
            framework = "Cambridge IGCSE"
        return code, name, framework

    def normalize_stage_name(self, sheet_name: str, subject_code: str) -> str:
        s = sheet_name.strip()
        if s == "Sheet6" and subject_code == "0058":
            return "Stage 6"
        if re.search(r"Stage\s*7\s*$", s, re.IGNORECASE):
            return "Stage 7"
        if re.search(r"Stage\s*8\s*$", s, re.IGNORECASE):
            return "Stage 8"
        if re.search(r"Stage\s*9\s*$", s, re.IGNORECASE):
            return "Stage 9"
        if re.search(r"Stage\s*1\s*$", s, re.IGNORECASE):
            return "Stage 1"
        if re.search(r"Stage\s*2\s*$", s, re.IGNORECASE):
            return "Stage 2"
        if re.search(r"Stage\s*3\s*$", s, re.IGNORECASE):
            return "Stage 3"
        if re.search(r"Stage\s*4\s*$", s, re.IGNORECASE):
            return "Stage 4"
        if re.search(r"Stage\s*5\s*$", s, re.IGNORECASE):
            return "Stage 5"
        if re.search(r"Stage\s*6\s*$", s, re.IGNORECASE):
            return "Stage 6"
        if "7-9" in s or "7 to 9" in s:
            return "Stages 7 to 9"
        if "1-6" in s or "1 to 6" in s:
            return "Stages 1 to 6"
        if "7&8" in s or "7 and 8" in s:
            return "Stages 7 and 8"
        if "3 and 4" in s or "3-4" in s:
            return "Stages 3 and 4"
        if "5 and 6" in s or "5-6" in s:
            return "Stages 5 and 6"
        return s

    def parse_sheet_rows(self, ws) -> List[List[str]]:
        rows = []
        for r in ws.iter_rows(values_only=True):
            cleaned_row = [clean_text(c) for c in r if c is not None and clean_text(c)]
            if cleaned_row:
                rows.append(cleaned_row)
        return rows

    def parse_workbook(self, file_path: Path) -> List[Dict[str, Any]]:
        code, subject_name, framework_name = self.extract_subject_info(file_path)
        wb = openpyxl.load_workbook(str(file_path), data_only=True, read_only=True)
        schemes = []

        for sheet_name in wb.sheetnames:
            if "Content" in sheet_name:
                continue

            ws = wb[sheet_name]
            rows = self.parse_sheet_rows(ws)
            if not rows:
                continue

            stage_name = self.normalize_stage_name(sheet_name, code)
            
            # Check if this sheet is tabular (Humanities style with Sub-strand in col 0 and LO in col 1)
            is_tabular = any(len(r) >= 2 and parse_lo_line(r[1]) is not None for r in rows[:15])

            topics_map: Dict[str, Dict[str, Any]] = {}
            current_topic_name = "General"
            current_subtopic_name = None

            if is_tabular:
                current_strand = sheet_name.replace("Stages 1 to 3", "").replace("Stages 4 to 6", "").replace("Stages 7 to 9", "").strip() or "General"
                
                for r in rows:
                    if len(r) == 1:
                        h = r[0]
                        if not parse_lo_line(h) and not h.lower().startswith("enquir") and not h.lower().startswith("learning"):
                            current_strand = h
                        continue

                    col0 = r[0]
                    col1 = r[1] if len(r) > 1 else ""
                    
                    lo_parsed = parse_lo_line(col1) or parse_lo_line(col0)
                    if lo_parsed:
                        lo_code, lo_text = lo_parsed
                        subtopic = col0 if parse_lo_line(col1) else None
                        topic_title = current_strand
                        
                        if topic_title not in topics_map:
                            topics_map[topic_title] = {"title": topic_title, "los": []}
                        
                        topics_map[topic_title]["los"].append({
                            "code": lo_code,
                            "text": lo_text,
                            "subtopic": subtopic,
                        })
            else:
                col_rows = []
                for r in rows:
                    col_rows.append(r[0])

                start_idx = 0
                for i, text in enumerate(col_rows):
                    if subject_name.lower() in text.lower() or "stage" in text.lower():
                        continue
                    start_idx = i
                    break

                i = start_idx
                while i < len(col_rows):
                    line = col_rows[i]

                    if "assessment objective" in line.lower() or "assessment objectives" in line.lower():
                        break

                    lo_parsed = parse_lo_line(line)
                    if lo_parsed:
                        lo_code, lo_text = lo_parsed
                        if current_topic_name not in topics_map:
                            topics_map[current_topic_name] = {"title": current_topic_name, "los": []}
                        topics_map[current_topic_name]["los"].append({
                            "code": lo_code,
                            "text": lo_text,
                            "subtopic": current_subtopic_name,
                        })
                        i += 1
                    else:
                        if i + 1 < len(col_rows) and parse_lo_line(col_rows[i + 1]) is None and not any(k in col_rows[i+1].lower() for k in ["assessment", "stage"]):
                            current_topic_name = line
                            current_subtopic_name = col_rows[i + 1]
                            if current_topic_name not in topics_map:
                                topics_map[current_topic_name] = {"title": current_topic_name, "los": []}
                            i += 2
                        else:
                            current_topic_name = line
                            current_subtopic_name = None
                            if current_topic_name not in topics_map:
                                topics_map[current_topic_name] = {"title": current_topic_name, "los": []}
                            i += 1

            topics_list = [t for t in topics_map.values() if t["los"]]
            if topics_list:
                scheme_seen_codes = set()
                for t in topics_list:
                    uniq = []
                    for lo in t["los"]:
                        base_code = lo["code"]
                        final_code = base_code
                        dup_counter = 1
                        while final_code in scheme_seen_codes:
                            dup_counter += 1
                            final_code = f"{base_code}.{dup_counter}"
                        scheme_seen_codes.add(final_code)
                        lo["code"] = final_code
                        uniq.append(lo)
                    t["los"] = uniq

                schemes.append({
                    "subject_code": code,
                    "subject_name": subject_name,
                    "framework": framework_name,
                    "year_group": stage_name,
                    "title": f"{subject_name} {stage_name}",
                    "topics": topics_list,
                })

        wb.close()
        return schemes

    def parse_igcse_pdf(self, file_path: Path) -> Optional[Dict[str, Any]]:
        filename = file_path.stem
        code, subject_name, framework_name = self.extract_subject_info(file_path)

        reader = pypdf.PdfReader(str(file_path))

        start_page = None
        end_page = None

        for p_idx in range(min(6, len(reader.pages))):
            text = reader.pages[p_idx].extract_text() or ""
            m_start = re.search(r"3\s+Subject\s+content\s*\.{2,}\s*(\d+)", text, re.IGNORECASE)
            if not m_start:
                m_start = re.search(r"3\s+Approaches\s+to\s+teaching\s*\.{2,}\s*(\d+)", text, re.IGNORECASE)
            m_end = re.search(r"4\s+Details\s+of\s+the\s+assessment\s*\.{2,}\s*(\d+)", text, re.IGNORECASE)
            if not m_end:
                m_end = re.search(r"4\s+Assessment\s+overview\s*\.{2,}\s*(\d+)", text, re.IGNORECASE)

            if m_start and start_page is None:
                start_page = int(m_start.group(1)) - 1
            if m_end and end_page is None:
                end_page = int(m_end.group(1)) - 1

        if start_page is None:
            start_page = 10
        if end_page is None or end_page <= start_page:
            end_page = min(len(reader.pages) - 4, start_page + 35)

        full_text = ""
        for p in range(start_page, end_page + 1):
            if p < len(reader.pages):
                full_text += "\n" + (reader.pages[p].extract_text() or "")

        lines = [clean_text(l) for l in full_text.split("\n") if clean_text(l)]

        topics: List[Dict[str, Any]] = []
        current_topic: Optional[Dict[str, Any]] = None
        current_subtopic: Optional[str] = None

        topic_header_re = re.compile(r"^(\d{1,2})\s+([A-Z][A-Za-z0-9\s,\-\(\)\/]{2,70})$")
        subtopic_header_re = re.compile(r"^(\d{1,2}\.\d{1,2})\s+([A-Z][A-Za-z0-9\s,\-\(\)\/]{2,80})$")
        lo_numbered_re = re.compile(r"^(\d{1,2}\.\d{1,2}\.\d{1,2})\s+(.*)")
        science_item_re = re.compile(r"^(\d{1,2})\s+([A-Z][a-z]+.*)")
        bullet_re = re.compile(r"^[\*\-\u2022]\s*(.*)")
        lo_action_re = re.compile(
            r"^(State|Describe|Explain|Understand|Know|Identify|Outline|Calculate|Demonstrate|Discuss|Evaluate|Analyse|Analyze|Distinguish|Compare|Recall|Use|Draw|Define|Recognise|Recognize|Select|Plan|Suggest|Show|Interpret|Apply)\s+(.*)",
            re.IGNORECASE,
        )

        if code == "0457":
            # Global Perspectives
            current_topic = {"title": "Global Topics & Issues", "los": []}
            gp_topics = [
                "Arts in society", "Change in culture and communities", "Climate change, energy and resources",
                "Conflict and peace", "Development, trade and aid", "Digital world", "Education for all",
                "Employment", "Environment, pollution and conservation", "Globalisation", "Health and wellbeing",
                "Law and criminality", "Media and communication", "Migration and urbanisation",
                "Political power and action", "Poverty and inequality", "Social identity and inclusion",
                "Sport and recreation", "Technology, industry and innovation", "Transport, travel and tourism",
                "Values and beliefs", "Water, food and agriculture"
            ]
            for idx, gpt in enumerate(gp_topics, 1):
                current_topic["los"].append({
                    "code": f"0457.T{idx:02d}",
                    "text": f"Explore perspectives, global and local issues in: {gpt}",
                    "subtopic": gpt,
                })
            topics.append(current_topic)

            skills_topic = {"title": "Global Perspectives Core Skills", "los": []}
            skills = [
                ("0457.S01", "Research: Research information and perspectives on global issues"),
                ("0457.S02", "Analysis: Analyse perspectives, causes and consequences of global issues"),
                ("0457.S03", "Evaluation: Evaluate sources, arguments, claims and lines of reasoning"),
                ("0457.S04", "Reflection: Reflect on own learning and how personal perspective has developed"),
                ("0457.S05", "Communication: Communicate arguments, evidence and reasoning clearly"),
                ("0457.S06", "Collaboration: Collaborate with others to design and complete a project"),
            ]
            for scode, stext in skills:
                skills_topic["los"].append({
                    "code": scode,
                    "text": stext,
                    "subtopic": "Core Skills",
                })
            topics.append(skills_topic)
        else:
            for line in lines:
                if any(k in line.lower() for k in ["cambridge international", "syllabus for", "back to contents", "subject content"]) or line.isdigit():
                    continue
                if line.lower() in ["core", "supplement", "core content", "extended content"]:
                    continue

                m_top = topic_header_re.match(line)
                if m_top and int(m_top.group(1)) <= 30:
                    if current_topic and current_topic["los"]:
                        topics.append(current_topic)
                    current_topic = {"title": f"{m_top.group(1)} {m_top.group(2).strip()}", "los": []}
                    current_subtopic = None
                    continue

                m_sub = subtopic_header_re.match(line)
                if m_sub:
                    current_subtopic = f"{m_sub.group(1)} {m_sub.group(2).strip()}"
                    if not current_topic:
                        current_topic = {"title": "General", "los": []}
                    continue

                m_lo_num = lo_numbered_re.match(line)
                if m_lo_num:
                    if not current_topic:
                        current_topic = {"title": "General", "los": []}
                    code_str = m_lo_num.group(1)
                    text_str = m_lo_num.group(2).strip()
                    current_topic["los"].append({
                        "code": f"{code}.{code_str}",
                        "text": text_str,
                        "subtopic": current_subtopic,
                    })
                    continue

                if code in ["0610", "0620", "0625"]:
                    m_sci = science_item_re.match(line)
                    if m_sci and len(line) > 15:
                        if not current_topic:
                            current_topic = {"title": "General", "los": []}
                        prefix = current_subtopic.split()[0] if current_subtopic else f"{code}.{len(topics)+1}"
                        current_topic["los"].append({
                            "code": f"{prefix}.{m_sci.group(1)}",
                            "text": m_sci.group(2).strip(),
                            "subtopic": current_subtopic,
                        })
                        continue

                m_bullet = bullet_re.match(line)
                m_action = lo_action_re.match(line)
                if (m_bullet or m_action) and len(line) > 15:
                    if not current_topic:
                        current_topic = {"title": "General", "los": []}
                    raw_text = m_bullet.group(1).strip() if m_bullet else line
                    prefix = current_subtopic.split()[0] if current_subtopic else f"{code}.{len(topics)+1}"
                    idx = len(current_topic["los"]) + 1
                    current_topic["los"].append({
                        "code": f"{prefix}.{idx}",
                        "text": raw_text,
                        "subtopic": current_subtopic,
                    })

            if current_topic and current_topic["los"]:
                topics.append(current_topic)

        # Deduplicate and guarantee unique LO codes scheme-wide
        scheme_seen_codes = set()
        for t in topics:
            uniq = []
            for lo in t["los"]:
                base_code = lo["code"]
                final_code = base_code
                dup_counter = 1
                while final_code in scheme_seen_codes:
                    dup_counter += 1
                    final_code = f"{base_code}.{dup_counter}"
                scheme_seen_codes.add(final_code)
                lo["code"] = final_code
                uniq.append(lo)
            t["los"] = uniq

        topics = [t for t in topics if t["los"]]
        if not topics:
            return None

        return {
            "subject_code": code,
            "subject_name": subject_name,
            "framework": "Cambridge IGCSE",
            "year_group": "Years 10-11 (IGCSE)",
            "title": f"{subject_name} IGCSE",
            "topics": topics,
        }

    def extract_all(self) -> Dict[str, Any]:
        workbooks = self.discover_workbooks()
        igcse_pdfs = self.discover_pdfs()

        frameworks_data: Dict[str, Dict[str, Any]] = {
            "Cambridge Primary": {
                "code": "CAMBRIDGE_PRIMARY",
                "name": "Cambridge Primary",
                "publisher": "Cambridge International",
                "schemes": [],
            },
            "Cambridge Lower Secondary": {
                "code": "CAMBRIDGE_LOWER_SECONDARY",
                "name": "Cambridge Lower Secondary",
                "publisher": "Cambridge International",
                "schemes": [],
            },
            "Cambridge IGCSE": {
                "code": "CAMBRIDGE_IGCSE",
                "name": "Cambridge IGCSE",
                "publisher": "Cambridge International",
                "schemes": [],
            },
        }

        for wb_path in workbooks:
            schemes = self.parse_workbook(wb_path)
            for scheme in schemes:
                fw = scheme["framework"]
                if fw in frameworks_data:
                    frameworks_data[fw]["schemes"].append(scheme)

        for pdf_path in igcse_pdfs:
            scheme = self.parse_igcse_pdf(pdf_path)
            if scheme:
                frameworks_data["Cambridge IGCSE"]["schemes"].append(scheme)

        return frameworks_data
