import re
from datetime import date
from pathlib import Path

import openpyxl
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from apps.curriculum.models import (
    CurriculumFramework,
    LearningObjective,
    SchemeOfWork,
    Subtopic,
    Topic,
)


class Command(BaseCommand):
    help = "Imports a curriculum framework from an Excel workbook."

    def add_arguments(self, parser):
        parser.add_argument("file", type=str, help="Path to the Excel file")
        parser.add_argument(
            "--framework", type=str, default="Cambridge Primary", help="Framework name"
        )
        parser.add_argument(
            "--dry-run", action="store_true", help="Parse and report without saving"
        )

    def handle(self, *args, **options):
        file_path = options["file"]
        dry_run = options["dry_run"]
        framework_name = options["framework"]

        if not Path(file_path).exists():
            raise CommandError(f"File not found: {file_path}")

        # Extract subject code and name from filename
        filename = Path(file_path).stem
        match = re.match(r"^(\d{4})\s+([A-Za-z\s\&]+?)(?:\s+Learning.*)?$", filename)
        if match:
            subject_code = match.group(1)
            subject_name = match.group(2).strip()
        else:
            parts = filename.split()
            subject_code = parts[0] if parts else "UNKNOWN"
            subject_name = " ".join(parts[1:3]) if len(parts) > 1 else filename

        self.stdout.write(self.style.MIGRATE_HEADING(f"Parsing {subject_code} {subject_name}"))

        if not dry_run:
            framework, _ = CurriculumFramework.objects.get_or_create(
                name=framework_name,
                defaults={"code": framework_name.upper().replace(" ", "_")},
            )
        else:
            framework = CurriculumFramework(
                name=framework_name, code=framework_name.upper().replace(" ", "_")
            )

        wb = openpyxl.load_workbook(file_path, data_only=True)
        # LO code must contain at least one digit to distinguish from multi-word topic names like "Managing Data"
        lo_pattern = re.compile(r"^(\*?[A-Za-z0-9\.\-]*\d+[A-Za-z0-9\.\-]*)\s+(.*)")

        def is_objective(text):
            return lo_pattern.match(text) is not None

        for sheet_name in wb.sheetnames:
            if "Content" in sheet_name or "Sheet" in sheet_name:
                continue

            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                if row[0]:
                    val = str(row[0]).strip()
                    if val:
                        rows.append(val)

            if not rows:
                continue

            year_group = sheet_name.strip()

            # Find start (skip header rows)
            start_idx = 0
            for i, r in enumerate(rows):
                if subject_name.lower() in r.lower() or year_group.lower() in r.lower():
                    continue
                start_idx = i
                break

            topics = []
            current_topic = None
            current_subtopic = None

            i = start_idx
            while i < len(rows):
                line = rows[i]

                # Skip Assessment Objectives for now
                if (
                    "assessment objective" in line.lower()
                    or "assessment objectives" in line.lower()
                ):
                    break

                if is_objective(line):
                    match = lo_pattern.match(line)
                    code = match.group(1)
                    text = match.group(2)
                    if current_topic:
                        current_topic["los"].append(
                            {
                                "subtopic": current_subtopic,
                                "code": code,
                                "text": text,
                            }
                        )
                    i += 1
                else:
                    if i + 1 < len(rows) and not is_objective(rows[i + 1]):
                        # Two headings in a row -> Topic, Subtopic
                        current_topic = {"title": line, "los": []}
                        topics.append(current_topic)
                        current_subtopic = rows[i + 1]
                        i += 2
                    else:
                        # One heading
                        if current_topic is None:
                            current_topic = {"title": line, "los": []}
                            topics.append(current_topic)
                            current_subtopic = None
                        elif current_subtopic is not None:
                            current_subtopic = line
                        else:
                            current_topic = {"title": line, "los": []}
                            topics.append(current_topic)
                            current_subtopic = None
                        i += 1

            topics = [t for t in topics if t["los"]]

            # Determine idempotency and version bumping
            self.sync_scheme(framework, subject_code, subject_name, year_group, topics, dry_run)

    def sync_scheme(
        self, framework, subject_code, subject_name, year_group, parsed_topics, dry_run
    ):
        try:
            latest_scheme = SchemeOfWork.objects.filter(
                framework__name=framework.name, subject_code=subject_code, year_group=year_group
            ).latest("version")
        except SchemeOfWork.DoesNotExist:
            latest_scheme = None

        if not latest_scheme:
            self.stdout.write(
                self.style.SUCCESS(
                    f"  {year_group}: New scheme (v1). Parsed {len(parsed_topics)} topics."
                )
            )
            for t in parsed_topics:
                subtopics_count = len(set(lo["subtopic"] for lo in t["los"] if lo["subtopic"]))
                self.stdout.write(
                    f"    - {t['title']} ({len(t['los'])} LOs, {subtopics_count} subtopics)"
                )
            if not dry_run:
                self.create_scheme_version(
                    framework, subject_code, subject_name, year_group, parsed_topics, 1
                )
            return

        # Compare LOs
        existing_los = {
            lo.code: lo for lo in LearningObjective.objects.filter(scheme=latest_scheme)
        }
        parsed_lo_codes = set()
        parsed_lo_data = {}
        for t in parsed_topics:
            for lo in t["los"]:
                parsed_lo_codes.add(lo["code"])
                parsed_lo_data[lo["code"]] = lo

        existing_codes = set(existing_los.keys())

        if parsed_lo_codes != existing_codes:
            added = parsed_lo_codes - existing_codes
            removed = existing_codes - parsed_lo_codes
            self.stdout.write(
                self.style.WARNING(
                    f"  {year_group}: Structural changes detected (Added: {len(added)}, "
                    f"Removed: {len(removed)}). Bumping to v{latest_scheme.version + 1}."
                )
            )
            if not dry_run:
                latest_scheme.retired_on = date.today()
                latest_scheme.is_active = False
                latest_scheme.save()
                self.create_scheme_version(
                    framework,
                    subject_code,
                    subject_name,
                    year_group,
                    parsed_topics,
                    latest_scheme.version + 1,
                )
        else:
            # Check for text updates
            text_changed = False
            for code in parsed_lo_codes:
                if existing_los[code].text != parsed_lo_data[code]["text"]:
                    text_changed = True
                    break

            if text_changed:
                self.stdout.write(
                    self.style.SUCCESS(
                        f"  {year_group}: Text updates detected. "
                        f"Updating v{latest_scheme.version} in-place."
                    )
                )
                if not dry_run:
                    with transaction.atomic():
                        # Wipe and recreate topics/los for simplicity since we keep the same scheme
                        Topic.objects.filter(scheme=latest_scheme).delete()
                        self.insert_hierarchy(latest_scheme, parsed_topics)
            else:
                self.stdout.write(f"  {year_group}: No changes detected. Skipping.")

    @transaction.atomic
    def create_scheme_version(
        self, framework, subject_code, subject_name, year_group, parsed_topics, version
    ):
        scheme = SchemeOfWork.objects.create(
            framework=framework,
            subject_code=subject_code,
            subject_name=subject_name,
            year_group=year_group,
            title=f"{subject_name} {year_group}",
            version=version,
            published_on=date.today(),
            is_active=True,
        )
        self.insert_hierarchy(scheme, parsed_topics)

    def insert_hierarchy(self, scheme, parsed_topics):
        topic_seq = 1
        seen_codes = set()
        for t_data in parsed_topics:
            topic = Topic.objects.create(scheme=scheme, title=t_data["title"], sequence=topic_seq)
            topic_seq += 1

            subtopics_cache = {}
            subtopic_seq = 1

            lo_seq = 1
            for lo_data in t_data["los"]:
                if lo_data["code"] in seen_codes:
                    self.stdout.write(
                        self.style.WARNING(
                            f"    Skipping duplicate code in {scheme.year_group}: {lo_data['code']}"
                        )
                    )
                    continue
                seen_codes.add(lo_data["code"])

                st_title = lo_data["subtopic"]
                st_obj = None
                if st_title:
                    if st_title not in subtopics_cache:
                        st_obj = Subtopic.objects.create(
                            topic=topic, title=st_title, sequence=subtopic_seq
                        )
                        subtopics_cache[st_title] = st_obj
                        subtopic_seq += 1
                    else:
                        st_obj = subtopics_cache[st_title]

                LearningObjective.objects.create(
                    scheme=scheme,
                    topic=topic,
                    subtopic=st_obj,
                    code=lo_data["code"],
                    text=lo_data["text"],
                    sequence=lo_seq,
                )
                lo_seq += 1
